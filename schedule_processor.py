"""
船期資料處理模組
處理不同船公司的船期表格式
"""

import pdfplumber
import pandas as pd
import openpyxl
import re
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO


def parse_cosco_pdf(file):
    """
    解析 COSCO 船期 PDF
    規則：
    - 只選擇 HPX2 服務
    - T/S Port 使用 POD 欄位
    - 日期格式：MM-DD
    """
    schedules = []
    
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            
            for table in tables:
                if not table or len(table) < 3:
                    continue
                
                # 找標題行
                header_idx = None
                for i, row in enumerate(table):
                    if row and 'Service' in str(row):
                        header_idx = i
                        break
                
                if header_idx is None:
                    continue
                
                # 處理數據
                i = header_idx + 1
                while i < len(table):
                    row1 = table[i]
                    row2 = table[i+1] if i+1 < len(table) else None
                    
                    if not row1 or len(row1) < 10:
                        i += 1
                        continue
                    
                    service = str(row1[1]).strip()
                    
                    # 只處理 HPX2
                    if service != 'HPX2':
                        i += 2
                        continue
                    
                    vessel = str(row1[2]).strip()
                    voyage = str(row1[3]).strip()
                    
                    # ETD
                    etd_raw = str(row1[6]).strip()
                    etd_match = re.search(r'2026-\s*(\d{2})-\s*(\d{2})', etd_raw)
                    etd = f"{etd_match.group(1)}-{etd_match.group(2)}" if etd_match else ""
                    
                    # T/S Port
                    ts_port = str(row1[8]).strip()
                    
                    # Transit Time
                    transit = str(row1[12]).strip() if len(row1) > 12 else ""
                    
                    # ETA
                    eta = ""
                    if row2 and len(row2) > 9:
                        eta_raw = str(row2[9]).strip()
                        eta_match = re.search(r'2026-\s*(\d{2})-\s*(\d{2})', eta_raw)
                        eta = f"{eta_match.group(1)}-{eta_match.group(2)}" if eta_match else ""
                    
                    if vessel and voyage and etd:
                        schedules.append({
                            'CARRIER': 'COSCO',
                            'Service': service,
                            'Vessel': vessel,
                            'Voyage': voyage,
                            'ETD': etd,
                            'ETA': eta if eta else "",
                            'Transit Time': transit,
                            'T/S Port': ts_port
                        })
                    
                    i += 2
    
    return schedules


def parse_one_pdf(file):
    """
    解析 ONE 船期 PDF
    規則：
    - Transit Time 只保留數字
    - 日期省略年份：MM-DD
    - Vessel 和 Voyage 需要分離
    - TRANSSHIPMENT 填入 T/S Port
    """
    schedules = []
    
    with pdfplumber.open(file) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            
            lines = text.split('\n')
            
            i = 0
            while i < len(lines):
                line = lines[i].strip()
                
                # 找到 Transit Time
                transit_match = re.match(r'^(\d+)\s+DAY\(S\)', line)
                if transit_match:
                    transit_time = transit_match.group(1)
                    
                    vessel = ""
                    voyage = ""
                    
                    # Vessel/Voyage 信息
                    if 'Vessel / Voyage' in line:
                        if i+1 < len(lines):
                            vessel_line = lines[i+1].strip()
                            vessel_match = re.match(r'^([A-Z\s]+?)(\s+\d+[A-Z]+)$', vessel_line)
                            if vessel_match:
                                vessel = vessel_match.group(1).strip()
                                voyage = vessel_match.group(2).strip()
                    
                    if not vessel:
                        for j in range(i, min(i+3, len(lines))):
                            check_line = lines[j].strip()
                            vessel_match = re.match(r'^([A-Z][A-Z\s]+?)(\s+\d+[A-Z]+)\s', check_line)
                            if vessel_match:
                                vessel = vessel_match.group(1).strip()
                                voyage = vessel_match.group(2).strip()
                                break
                    
                    # 找日期和其他信息
                    etd = ""
                    eta = ""
                    service = ""
                    ts_port = ""
                    
                    for j in range(i, min(i+20, len(lines))):
                        check_line = lines[j].strip()
                        
                        # Origin Destination
                        if check_line == 'Origin Destination':
                            if j+1 < len(lines):
                                date_line = lines[j+1].strip()
                                dates = re.findall(r'(\d{4})-(\d{2})-(\d{2})', date_line)
                                if len(dates) >= 2:
                                    etd = f"{dates[0][1]}-{dates[0][2]}"
                                    eta = f"{dates[1][1]}-{dates[1][2]}"
                        
                        # TRANSSHIPMENT
                        if check_line == 'TRANSSHIPMENT':
                            ts_port = 'TRANSSHIPMENT'
                        
                        # Service
                        service_match = re.search(r'Service.*?Origin.*?Destination', check_line)
                        if service_match and j+1 < len(lines):
                            service_line = lines[j+1].strip()
                            parts = service_line.split()
                            if parts:
                                candidate = parts[0]
                                if candidate not in ['CY', 'Origin', 'Destination'] and len(candidate) <= 5:
                                    service = candidate
                    
                    if vessel and voyage and etd:
                        schedules.append({
                            'CARRIER': 'ONE',
                            'Service': service,
                            'Vessel': vessel,
                            'Voyage': voyage,
                            'ETD': etd,
                            'ETA': eta,
                            'Transit Time': transit_time,
                            'T/S Port': ts_port
                        })
                
                i += 1
    
    return schedules


def parse_sitc_excel(file):
    """
    解析 SITC 船期 Excel
    規則：
    - Service: 從 A1 擷取英文字母組合（排除符號）
    - T/S Port: 從 A1 抓取 "DIRECT" 字樣
    - Transit Time: 從 Row 4 抓取 "days" 前的數字
    - 過濾: Vessel 或 ETA 包含 "SKIP" 的記錄
    """
    schedules = []
    
    # 使用 data_only=True 讀取公式計算後的值
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    
    # 1. 從 A1 擷取 Service (只要英文字母和數字組合)
    a1_value = str(ws['A1'].value) if ws['A1'].value else ""
    service_match = re.search(r'([A-Z]+\d+[A-Z]*)', a1_value)
    service = service_match.group(1) if service_match else ""
    
    # 2. 從 A1 判斷是否為 DIRECT
    ts_port = "DIRECT" if "DIRECT" in a1_value.upper() else ""
    
    # 3. 從 Row 4 提取 Transit Time
    transit_time = ""
    row4_col4 = str(ws.cell(row=4, column=4).value) if ws.cell(row=4, column=4).value else ""
    transit_match = re.search(r'(\d+)\s*days?', row4_col4, re.IGNORECASE)
    if transit_match:
        transit_time = transit_match.group(1)
    
    # 4. 讀取數據（從 Row 5 開始）
    for row_idx in range(5, ws.max_row + 1):
        vessel = str(ws.cell(row=row_idx, column=1).value or "").strip()
        voyage = str(ws.cell(row=row_idx, column=2).value or "").strip()
        etd_raw = ws.cell(row=row_idx, column=3).value
        eta_raw = ws.cell(row=row_idx, column=4).value
        
        # 過濾條件: Vessel 或 ETA 包含 "SKIP"
        if "SKIP" in vessel.upper():
            continue
        
        if eta_raw and "SKIP" in str(eta_raw).upper():
            continue
        
        # 必須有 Vessel 和 Voyage
        if not vessel or not voyage:
            continue
        
        # 處理 ETD
        etd = ""
        if etd_raw:
            if isinstance(etd_raw, datetime):
                etd = etd_raw.strftime("%m-%d")
            elif isinstance(etd_raw, (int, float)):
                # Excel 日期序列號
                base_date = datetime(1899, 12, 30)
                date_value = base_date + timedelta(days=float(etd_raw))
                etd = date_value.strftime("%m-%d")
            else:
                date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(etd_raw))
                if date_match:
                    etd = f"{date_match.group(2)}-{date_match.group(3)}"
        
        # 處理 ETA
        eta = ""
        if eta_raw:
            if isinstance(eta_raw, datetime):
                eta = eta_raw.strftime("%m-%d")
            elif isinstance(eta_raw, (int, float)):
                base_date = datetime(1899, 12, 30)
                date_value = base_date + timedelta(days=float(eta_raw))
                eta = date_value.strftime("%m-%d")
            else:
                date_match = re.search(r'(\d{4})-(\d{2})-(\d{2})', str(eta_raw))
                if date_match:
                    eta = f"{date_match.group(2)}-{date_match.group(3)}"
        
        if vessel and voyage and etd:
            schedules.append({
                'CARRIER': 'SITC',
                'Service': service,
                'Vessel': vessel,
                'Voyage': voyage,
                'ETD': etd,
                'ETA': eta,
                'Transit Time': transit_time,
                'T/S Port': ts_port
            })
    
    return schedules


def process_schedules(files_data, carrier_mapping, remove_duplicates=True):
    """
    處理所有上傳的船期檔案
    
    Args:
        files_data: 檔案資料列表
        carrier_mapping: 檔案名稱 -> 船公司對應
        remove_duplicates: 是否移除重複記錄
    
    Returns:
        DataFrame: 處理後的船期資料
    """
    all_schedules = []
    
    for file_data in files_data:
        file_name = file_data['name']
        file_content = file_data['content']
        carrier = carrier_mapping.get(file_name, '自動識別')
        
        # 根據船公司類型和檔案格式選擇解析器
        is_excel = file_name.lower().endswith(('.xlsx', '.xls'))
        is_pdf = file_name.lower().endswith('.pdf')
        
        schedules = []
        
        try:
            if carrier == 'SITC' or 'SITC' in file_name.upper():
                # SITC 使用 Excel 格式
                if is_excel:
                    schedules = parse_sitc_excel(file_content)
                else:
                    # 如果是 PDF，可能需要轉換或提示錯誤
                    print(f"警告: SITC 通常使用 Excel 格式，檔案 {file_name} 是 PDF")
            elif carrier == 'COSCO' or 'COSCO' in file_name.upper():
                if is_pdf:
                    schedules = parse_cosco_pdf(file_content)
            elif carrier == 'ONE' or 'ONE' in file_name.upper():
                if is_pdf:
                    schedules = parse_one_pdf(file_content)
            else:
                # 嘗試自動識別
                if is_pdf:
                    # 先嘗試 COSCO
                    try:
                        schedules = parse_cosco_pdf(file_content)
                    except:
                        pass
                    
                    # 如果沒有結果，嘗試 ONE
                    if not schedules:
                        try:
                            schedules = parse_one_pdf(file_content)
                        except:
                            pass
                elif is_excel:
                    # Excel 可能是 SITC
                    try:
                        schedules = parse_sitc_excel(file_content)
                    except:
                        pass
        except Exception as e:
            print(f"處理檔案 {file_name} 時發生錯誤: {str(e)}")
            schedules = []
        
        all_schedules.extend(schedules)
    
    # 創建 DataFrame
    df = pd.DataFrame(all_schedules)
    
    if df.empty:
        return df
    
    # 移除重複
    if remove_duplicates:
        df = df.drop_duplicates()
    
    # 按 ETD 排序
    df = df.sort_values('ETD').reset_index(drop=True)
    
    return df


def create_excel_file(df, include_summary=True):
    """
    創建格式化的 Excel 檔案
    
    Args:
        df: 船期資料 DataFrame
        include_summary: 是否包含統計摘要工作表
    
    Returns:
        BytesIO: Excel 檔案的二進制內容
    """
    output = BytesIO()
    
    wb = Workbook()
    
    # 主要船期表
    ws = wb.active
    ws.title = "船期排序表"
    
    # 標題行
    headers = ['CARRIER', 'Service', 'Vessel', 'Voyage', 'ETD', 'ETA', 'Transit Time', 'T/S Port']
    ws.append(headers)
    
    # 格式化標題
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11, name='Arial')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_num in range(1, 9):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # 添加數據
    for _, row in df.iterrows():
        ws.append([
            row['CARRIER'],
            row['Service'] if pd.notna(row['Service']) else "",
            row['Vessel'],
            row['Voyage'],
            row['ETD'],
            row['ETA'] if pd.notna(row['ETA']) else "",
            row['Transit Time'] if pd.notna(row['Transit Time']) else "",
            row['T/S Port'] if pd.notna(row['T/S Port']) else ""
        ])
    
    # 設定列寬
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 15
    
    # 邊框和格式
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    data_font = Font(name='Arial', size=10)
    center_align = Alignment(horizontal='center', vertical='center')
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.font = data_font
                cell.alignment = center_align
    
    # 凍結首行
    ws.freeze_panes = 'A2'
    
    # 統計摘要工作表
    if include_summary and not df.empty:
        ws_summary = wb.create_sheet("統計摘要")
        
        # 標題
        ws_summary['A1'] = '船期統計摘要'
        ws_summary['A1'].font = Font(bold=True, size=14, name='Arial')
        ws_summary.merge_cells('A1:D1')
        
        # 基本統計
        ws_summary['A3'] = '項目'
        ws_summary['B3'] = '數量'
        ws_summary['A3'].font = Font(bold=True)
        ws_summary['B3'].font = Font(bold=True)
        
        row = 4
        ws_summary[f'A{row}'] = '總船期數'
        ws_summary[f'B{row}'] = len(df)
        
        # 各船公司統計
        row += 1
        ws_summary[f'A{row}'] = '船公司分佈'
        ws_summary[f'A{row}'].font = Font(bold=True)
        
        for carrier, count in df['CARRIER'].value_counts().items():
            row += 1
            ws_summary[f'A{row}'] = f'  - {carrier}'
            ws_summary[f'B{row}'] = count
        
        # 日期範圍
        row += 2
        ws_summary[f'A{row}'] = '日期範圍'
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
        ws_summary[f'A{row}'] = '  最早 ETD'
        ws_summary[f'B{row}'] = df['ETD'].min()
        row += 1
        ws_summary[f'A{row}'] = '  最晚 ETD'
        ws_summary[f'B{row}'] = df['ETD'].max()
        
        # 格式化摘要表
        for row in ws_summary.iter_rows(min_row=3, max_row=ws_summary.max_row, min_col=1, max_col=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 15
    
    wb.save(output)
    output.seek(0)
    
    return output


def get_statistics(df):
    """
    獲取船期統計資訊
    
    Returns:
        dict: 統計資訊字典
    """
    if df.empty:
        return {}
    
    stats = {
        'total': len(df),
        'carriers': df['CARRIER'].value_counts().to_dict(),
        'services': df['Service'].value_counts().to_dict(),
        'date_range': {
            'earliest': df['ETD'].min(),
            'latest': df['ETD'].max()
        },
        'ts_ports': df[df['T/S Port'] != '']['T/S Port'].value_counts().to_dict()
    }
    
    return stats
